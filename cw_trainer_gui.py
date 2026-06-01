"""
+===========================================================+
|           CW ADAPTIVE TRAINER                              |
|  Treinamento adaptativo de telegrafia                      |
|                                                            |
|  Dependencias: pip install openpyxl                        |
|  GUI: tkinter (nativo do Python)                           |
|  Audio: winsound (nativo do Windows)                       |
+===========================================================+
"""

import os
import sys
import json
import math
import wave
import array
import random
import time
import datetime
import tempfile
import threading
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext

if sys.platform == 'win32':
    import winsound
else:
    winsound = None

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("openpyxl necessário: pip install openpyxl")
    sys.exit(1)


# ===========================================================
#  CONSTANTES
# ===========================================================

CHAR_SEQUENCE = list("ETIANMSURWDKGOHVFLPJBXCYZQ0123456789.,?/=")

MORSE_CODE = {
    'A': '.-',    'B': '-...',  'C': '-.-.',  'D': '-..',   'E': '.',
    'F': '..-.',  'G': '--.',   'H': '....',  'I': '..',    'J': '.---',
    'K': '-.-',   'L': '.-..',  'M': '--',    'N': '-.',    'O': '---',
    'P': '.--.',  'Q': '--.-',  'R': '.-.',   'S': '...',   'T': '-',
    'U': '..-',   'V': '...-',  'W': '.--',   'X': '-..-',  'Y': '-.--',
    'Z': '--..',
    '0': '-----', '1': '.----', '2': '..---', '3': '...--', '4': '....-',
    '5': '.....', '6': '-....', '7': '--...', '8': '---..', '9': '----.',
    '.': '.-.-.-', ',': '--..--', '?': '..--..', '/': '-..-.', '=': '-...-',
}

SAMPLE_RATE = 22050
RAMP_MS = 3
TWO_PI = 2.0 * math.pi


def get_data_path(filename):
    if getattr(sys, 'frozen', False):
        base = os.path.dirname(sys.executable)
    else:
        base = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, filename)


CONFIG_FILE = get_data_path("cw_config.json")
EXCEL_FILE = get_data_path("cw_telemetria.xlsx")


# ===========================================================
#  AUDIO ENGINE
# ===========================================================

def generate_tone_fast(duration_s, frequency):
    n = int(SAMPLE_RATE * duration_s)
    if n == 0:
        return array.array('h')
    ramp = int(SAMPLE_RATE * RAMP_MS / 1000)
    buf = array.array('h', [0] * n)
    amp = 24000
    for i in range(n):
        val = math.sin(TWO_PI * frequency * i / SAMPLE_RATE)
        if i < ramp:
            val *= i / ramp
        elif i >= n - ramp:
            val *= (n - 1 - i) / ramp
        buf[i] = int(val * amp)
    return buf


def generate_silence_fast(duration_s):
    return array.array('h', [0] * int(SAMPLE_RATE * duration_s))


def calculate_timing(wpm, farnsworth_wpm):
    dit = 1.2 / wpm
    dah = 3 * dit
    intra_char = dit
    if farnsworth_wpm < wpm:
        t_char = 50 * dit
        t_farn = 60.0 / farnsworth_wpm
        delta = (t_farn - t_char) / 19.0
        inter_char = 3 * delta
        inter_word = 7 * delta
    else:
        inter_char = 3 * dit
        inter_word = 7 * dit
    return dit, dah, intra_char, inter_char, inter_word


def text_to_wav(text, frequency, wpm, farnsworth_wpm):
    dit_t, dah_t, intra_t, inter_char_t, inter_word_t = calculate_timing(wpm, farnsworth_wpm)
    tone_dit = generate_tone_fast(dit_t, frequency)
    tone_dah = generate_tone_fast(dah_t, frequency)
    sil_intra = generate_silence_fast(intra_t)
    sil_char = generate_silence_fast(inter_char_t)
    sil_word = generate_silence_fast(inter_word_t)
    audio = array.array('h')
    for i, char in enumerate(text):
        if char == ' ':
            audio.extend(sil_word)
            continue
        morse = MORSE_CODE.get(char.upper(), '')
        if not morse:
            continue
        for j, symbol in enumerate(morse):
            if symbol == '.':
                audio.extend(tone_dit)
            elif symbol == '-':
                audio.extend(tone_dah)
            if j < len(morse) - 1:
                audio.extend(sil_intra)
        if i < len(text) - 1 and text[i + 1] != ' ':
            audio.extend(sil_char)
    return audio


def save_wav(audio_array, filepath):
    with wave.open(filepath, 'w') as wf:
        wf.setnchannels(1)
        wf.setsampwidth(2)
        wf.setframerate(SAMPLE_RATE)
        wf.writeframes(audio_array.tobytes())


def play_wav_blocking(filepath):
    if sys.platform == 'win32' and winsound:
        winsound.PlaySound(filepath, winsound.SND_FILENAME)
    elif sys.platform == 'darwin':
        os.system(f'afplay "{filepath}"')
    else:
        os.system(f'aplay -q "{filepath}" 2>/dev/null || paplay "{filepath}" 2>/dev/null')


def stop_audio():
    if sys.platform == 'win32' and winsound:
        winsound.PlaySound(None, winsound.SND_PURGE)


# ===========================================================
#  CONFIG & PERSISTENCE
# ===========================================================

DEFAULT_CONFIG = {
    "frequency": 700,
    "freq_variation": 0,
    "wpm": 10,
    "farnsworth": 6,
    "lesson": 2,
    "num_groups": 6,
    "group_size": 4,
    "pass_threshold": 90,
    "window_size": 10,
    "weight_floor": 0.1,
    "weight_ceiling": 1.0,
    "char_history": {},
}


def load_config():
    config = dict(DEFAULT_CONFIG)
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                saved = json.load(f)
            config.update(saved)
        except (json.JSONDecodeError, IOError):
            pass
    if not isinstance(config.get("char_history"), dict):
        config["char_history"] = {}
    return config


def save_config(config):
    with open(CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=2)


def get_char_accuracy(config, char):
    history = config["char_history"].get(char, [])
    if not history:
        return 0.5
    return sum(history) / len(history)


def update_char_history(config, char, hit):
    window = config.get("window_size", 10)
    if char not in config["char_history"]:
        config["char_history"][char] = []
    config["char_history"][char].append(1 if hit else 0)
    if len(config["char_history"][char]) > window:
        config["char_history"][char] = config["char_history"][char][-window:]


# ===========================================================
#  WEIGHTED GROUP GENERATION
# ===========================================================

def calculate_weights(config):
    floor = config.get("weight_floor", 0.1)
    ceiling = config.get("weight_ceiling", 1.0)
    lesson_chars = CHAR_SEQUENCE[:config["lesson"]]
    weights = []
    for char in lesson_chars:
        acc = get_char_accuracy(config, char)
        w = floor + (ceiling - floor) * (1.0 - acc)
        weights.append(w)
    return lesson_chars, weights


def generate_groups_weighted(config):
    lesson_chars, weights = calculate_weights(config)
    num_groups = config.get("num_groups", 6)
    group_size = config.get("group_size", 4)
    groups = []
    for _ in range(num_groups):
        group = ''.join(random.choices(lesson_chars, weights=weights, k=group_size))
        groups.append(group)
    return groups


def randomize_frequency(config):
    base = config["frequency"]
    var = config.get("freq_variation", 0)
    if var == 0:
        return base
    return base + random.randint(-var, var)


# ===========================================================
#  COMPARISON
# ===========================================================

def compare_groups(sent_groups, received_text):
    recv_groups = received_text.upper().split()
    total = correct = 0
    details = []
    char_results = []
    for i, sg in enumerate(sent_groups):
        rg = recv_groups[i] if i < len(recv_groups) else ''
        gc = 0
        for ci in range(len(sg)):
            total += 1
            hit = ci < len(rg) and sg[ci] == rg[ci]
            if hit:
                gc += 1
            char_results.append((sg[ci], hit))
        correct += gc
        details.append((sg, rg, gc))
    accuracy = (correct / total * 100) if total > 0 else 0.0
    return total, correct, accuracy, details, char_results


# ===========================================================
#  EXCEL TELEMETRY
# ===========================================================

def init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetria CW"
    headers = ["Data", "Hora", "Lição", "Caracteres", "WPM", "Farnsworth",
               "Frequência (Hz)", "Grupos", "Chars/Grupo", "Total Chars",
               "Chars Corretos", "Acerto (%)", "Resultado"]
    hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill
        c.font = hfont
        c.alignment = Alignment(horizontal='center')
        c.border = tb
    for letter, w in {'A':12,'B':10,'C':6,'D':40,'E':6,'F':12,'G':15,
                       'H':8,'I':11,'J':12,'K':14,'L':10,'M':12}.items():
        ws.column_dimensions[letter].width = w
    wb.save(EXCEL_FILE)


def log_to_excel(config, lesson_chars, freq, total, correct, accuracy, passed):
    init_excel()
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    now = datetime.datetime.now()
    res = "APROVADO" if passed else "REPETIR"
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    row = [now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), config["lesson"],
           ' '.join(lesson_chars), config["wpm"], config["farnsworth"], freq,
           config["num_groups"], config["group_size"], total, correct,
           round(accuracy, 1), res]
    ws.append(row)
    rn = ws.max_row
    for col in range(1, len(row) + 1):
        c = ws.cell(row=rn, column=col)
        c.border = tb
        c.alignment = Alignment(horizontal='center')
    rc = ws.cell(row=rn, column=len(row))
    if res == "APROVADO":
        rc.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        rc.font = Font(color="006100", bold=True)
    else:
        rc.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        rc.font = Font(color="9C0006", bold=True)
    wb.save(EXCEL_FILE)


# ===========================================================
#  GUI APPLICATION
# ===========================================================

class CWTrainerApp:
    BG = "#1e1e2e"
    BG2 = "#282840"
    BG3 = "#313150"
    FG = "#e0e0e0"
    FG2 = "#a0a0b8"
    ACCENT = "#f0c040"
    GREEN = "#40c060"
    RED = "#e04050"

    def __init__(self, root):
        self.root = root
        self.root.title("Treino de CW por PY2TAE (Lucas)")
        self.root.geometry("880x740")
        self.root.minsize(780, 650)
        self.root.configure(bg=self.BG)
        self.config = load_config()
        init_excel()
        self.current_groups = []
        self.current_freq = 0
        self.is_playing = False
        self.exercise_active = False
        self._setup_styles()
        self._build_ui()
        self._update_info()

    def _setup_styles(self):
        s = ttk.Style()
        s.theme_use('clam')
        s.configure("Main.TFrame", background=self.BG)
        s.configure("Card.TFrame", background=self.BG2)
        s.configure("Main.TLabel", background=self.BG, foreground=self.FG, font=("Consolas", 10))
        s.configure("Title.TLabel", background=self.BG, foreground=self.ACCENT, font=("Consolas", 14, "bold"))
        s.configure("Header.TLabel", background=self.BG2, foreground=self.ACCENT, font=("Consolas", 10, "bold"))
        s.configure("Status.TLabel", background=self.BG3, foreground=self.FG2, font=("Consolas", 9))
        s.configure("Score.TLabel", background=self.BG2, foreground=self.GREEN, font=("Consolas", 32, "bold"))
        s.configure("Play.TButton", font=("Consolas", 11, "bold"), padding=8)
        s.configure("Nav.TButton", font=("Consolas", 10), padding=6)
        s.configure("Main.TNotebook", background=self.BG)
        s.configure("Main.TNotebook.Tab", font=("Consolas", 10, "bold"), padding=[14, 5])

    def _build_ui(self):
        self.nb = ttk.Notebook(self.root, style="Main.TNotebook")
        self.nb.pack(fill=tk.BOTH, expand=True, padx=6, pady=(6, 0))
        self.tab_ex = ttk.Frame(self.nb, style="Main.TFrame")
        self.nb.add(self.tab_ex, text="  Exercício  ")
        self._build_exercise()
        self.tab_cfg = ttk.Frame(self.nb, style="Main.TFrame")
        self.nb.add(self.tab_cfg, text="  Config  ")
        self._build_config()
        self.tab_st = ttk.Frame(self.nb, style="Main.TFrame")
        self.nb.add(self.tab_st, text="  Estatísticas  ")
        self._build_stats()
        self.status_var = tk.StringVar(value="Pronto")
        ttk.Label(self.root, textvariable=self.status_var, style="Status.TLabel",
                  anchor=tk.W, padding=(10, 4)).pack(fill=tk.X, side=tk.BOTTOM)

    # ── Exercise Tab ───────────────────────────────────────

    def _build_exercise(self):
        t = self.tab_ex
        info = ttk.Frame(t, style="Card.TFrame")
        info.pack(fill=tk.X, padx=10, pady=(10, 5))
        self.info_var = tk.StringVar()
        ttk.Label(info, textvariable=self.info_var, style="Header.TLabel",
                  wraplength=820).pack(padx=10, pady=8, anchor=tk.W)

        bf = ttk.Frame(t, style="Main.TFrame")
        bf.pack(fill=tk.X, padx=10, pady=5)
        self.btn_play = ttk.Button(bf, text="PLAY", style="Play.TButton", command=self._on_play)
        self.btn_play.pack(side=tk.LEFT, padx=(0, 5))
        self.btn_replay = ttk.Button(bf, text="Repetir", style="Nav.TButton",
                                     command=self._on_replay, state=tk.DISABLED)
        self.btn_replay.pack(side=tk.LEFT, padx=5)
        self.btn_sample = ttk.Button(bf, text="Amostra", style="Nav.TButton", command=self._on_sample)
        self.btn_sample.pack(side=tk.LEFT, padx=5)
        self.btn_check = ttk.Button(bf, text="Corrigir", style="Play.TButton",
                                    command=self._on_check, state=tk.DISABLED)
        self.btn_check.pack(side=tk.RIGHT, padx=(5, 0))

        sf = ttk.Frame(t, style="Card.TFrame")
        sf.pack(fill=tk.X, padx=10, pady=5)
        self.score_var = tk.StringVar(value="--")
        ttk.Label(sf, textvariable=self.score_var, style="Score.TLabel").pack(pady=4)

        ttk.Label(t, text="Digite o que ouviu (separe grupos com espaço):",
                  style="Main.TLabel").pack(anchor=tk.W, padx=10, pady=(6, 2))
        self.inp = tk.Text(t, height=3, font=("Consolas", 16), bg=self.BG2,
                           fg=self.ACCENT, insertbackground=self.ACCENT,
                           relief=tk.FLAT, padx=10, pady=8, wrap=tk.WORD)
        self.inp.pack(fill=tk.X, padx=10, pady=(0, 5))
        self.inp.bind("<Return>", lambda e: (self._on_check(), "break"))

        ttk.Label(t, text="Resultado:", style="Main.TLabel").pack(anchor=tk.W, padx=10, pady=(4, 2))
        self.log = scrolledtext.ScrolledText(t, height=12, font=("Consolas", 10),
                                              bg=self.BG3, fg=self.FG, relief=tk.FLAT,
                                              padx=8, pady=6, state=tk.DISABLED)
        self.log.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        for tag, color in [("ok", self.GREEN), ("err", self.RED),
                           ("accent", self.ACCENT), ("info", self.FG2),
                           ("bold", self.FG)]:
            kw = {"foreground": color}
            if tag == "bold":
                kw["font"] = ("Consolas", 10, "bold")
            self.log.tag_configure(tag, **kw)

    # ── Config Tab ─────────────────────────────────────────

    def _build_config(self):
        t = self.tab_cfg
        canvas = tk.Canvas(t, bg=self.BG, highlightthickness=0)
        sb = ttk.Scrollbar(t, orient=tk.VERTICAL, command=canvas.yview)
        sf = ttk.Frame(canvas, style="Main.TFrame")
        sf.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=sf, anchor="nw")
        canvas.configure(yscrollcommand=sb.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        ttk.Label(sf, text="CONFIGURAÇÕES", style="Title.TLabel").pack(anchor=tk.W, pady=(0, 10))

        self.cfg_vars = {}
        fields = [
            ("lesson",         "Lição (2-41)",               "int",   2, 41),
            ("frequency",      "Frequência base (Hz)",       "int",   200, 1500),
            ("freq_variation", "Variação +/- (Hz)",          "int",   0, 200),
            ("wpm",            "WPM (velocidade caractere)", "int",   5, 60),
            ("farnsworth",     "Farnsworth WPM",             "int",   2, 60),
            ("num_groups",     "Grupos por exercício",       "int",   1, 50),
            ("group_size",     "Caracteres por grupo",       "int",   2, 10),
            ("pass_threshold", "Meta de acerto (%)",         "float", 50, 100),
            ("window_size",    "Janela média móvel",         "int",   3, 100),
            ("weight_floor",   "Peso mínimo (char bom)",     "float", 0.01, 0.5),
            ("weight_ceiling", "Peso máximo (char ruim)",    "float", 0.1, 10.0),
        ]
        for key, label, typ, vmin, vmax in fields:
            row = ttk.Frame(sf, style="Main.TFrame")
            row.pack(fill=tk.X, pady=3)
            ttk.Label(row, text=label, style="Main.TLabel", width=28).pack(side=tk.LEFT)
            var = tk.StringVar(value=str(self.config.get(key, DEFAULT_CONFIG[key])))
            self.cfg_vars[key] = (var, typ, vmin, vmax)
            tk.Entry(row, textvariable=var, font=("Consolas", 11), width=10,
                     bg=self.BG2, fg=self.ACCENT, insertbackground=self.ACCENT,
                     relief=tk.FLAT).pack(side=tk.LEFT, padx=10)
            ttk.Label(row, text=f"({vmin}-{vmax})", style="Status.TLabel").pack(side=tk.LEFT)

        bf = ttk.Frame(sf, style="Main.TFrame")
        bf.pack(fill=tk.X, pady=15)
        ttk.Button(bf, text="Salvar", style="Play.TButton", command=self._save_cfg).pack(side=tk.LEFT, padx=(0,10))
        ttk.Button(bf, text="Resetar Stats", style="Nav.TButton", command=self._reset).pack(side=tk.LEFT, padx=10)
        ttk.Button(bf, text="Restaurar Padrões", style="Nav.TButton", command=self._defaults).pack(side=tk.LEFT, padx=10)
        self.cfg_info = tk.StringVar()
        ttk.Label(sf, textvariable=self.cfg_info, style="Main.TLabel", wraplength=700).pack(anchor=tk.W, pady=5)

    # ── Stats Tab ──────────────────────────────────────────

    def _build_stats(self):
        t = self.tab_st
        bf = ttk.Frame(t, style="Main.TFrame")
        bf.pack(fill=tk.X, padx=10, pady=10)
        ttk.Label(bf, text="ESTATÍSTICAS POR CARACTERE", style="Title.TLabel").pack(side=tk.LEFT)
        ttk.Button(bf, text="Atualizar", style="Nav.TButton", command=self._refresh_stats).pack(side=tk.RIGHT)
        self.st_txt = scrolledtext.ScrolledText(t, font=("Consolas", 10), bg=self.BG3, fg=self.FG,
                                                 relief=tk.FLAT, padx=8, pady=6, state=tk.DISABLED)
        self.st_txt.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        for tag, color in [("ok", self.GREEN), ("err", self.RED),
                           ("accent", self.ACCENT), ("info", self.FG2),
                           ("header", self.ACCENT)]:
            kw = {"foreground": color}
            if tag == "header":
                kw["font"] = ("Consolas", 10, "bold")
            self.st_txt.tag_configure(tag, **kw)
        self._refresh_stats()

    # ── Helpers ────────────────────────────────────────────

    def _l(self, text, tag=None):
        self.log.config(state=tk.NORMAL)
        self.log.insert(tk.END, text, tag) if tag else self.log.insert(tk.END, text)
        self.log.see(tk.END)
        self.log.config(state=tk.DISABLED)

    def _lc(self):
        self.log.config(state=tk.NORMAL)
        self.log.delete("1.0", tk.END)
        self.log.config(state=tk.DISABLED)

    def _update_info(self):
        c = self.config
        lc = CHAR_SEQUENCE[:c["lesson"]]
        txt = (f"Lição {c['lesson']}/{len(CHAR_SEQUENCE)}  |  "
               f"{c['wpm']} WPM  Farns: {c['farnsworth']}  |  "
               f"{c['frequency']}Hz +/-{c['freq_variation']}  |  "
               f"{c['num_groups']}x{c['group_size']}  |  "
               f"Meta: {c['pass_threshold']}%  |  Janela: {c['window_size']}\n"
               f"Chars: {' '.join(lc)}")
        if c["lesson"] < len(CHAR_SEQUENCE):
            nx = CHAR_SEQUENCE[c["lesson"]]
            txt += f"  |  Próximo: {nx} ({MORSE_CODE.get(nx, '')})"
        self.info_var.set(txt)
        # Status bar
        cwd = [ch for ch in lc if len(c["char_history"].get(ch, [])) >= c["window_size"]]
        if cwd:
            accs = [get_char_accuracy(c, ch) for ch in cwd]
            avg = sum(accs) / len(accs) * 100
            below = sum(1 for a in accs if a < c["pass_threshold"] / 100.0)
            self.status_var.set(f"Lição {c['lesson']}  |  Media: {avg:.0f}%  |  Abaixo da meta: {below}/{len(lc)}")
        else:
            self.status_var.set(f"Lição {c['lesson']}  |  Sem dados. Comece um exercício!")

    def _set_buttons(self, play=True, check=False, replay=False, sample=True):
        self.btn_play.config(state=tk.NORMAL if play else tk.DISABLED)
        self.btn_check.config(state=tk.NORMAL if check else tk.DISABLED)
        self.btn_replay.config(state=tk.NORMAL if replay else tk.DISABLED)
        self.btn_sample.config(state=tk.NORMAL if sample else tk.DISABLED)

    # ── Exercise Actions ───────────────────────────────────

    def _on_play(self):
        if self.is_playing:
            return
        self._lc()
        self.score_var.set("--")
        self.inp.delete("1.0", tk.END)
        self.current_groups = generate_groups_weighted(self.config)
        self.current_freq = randomize_frequency(self.config)
        text = ' '.join(self.current_groups)
        total = len(text.replace(' ', ''))
        self._l(f"Exercicio: {self.config['num_groups']}x{self.config['group_size']} = {total} chars  |  {self.current_freq} Hz\n", "info")
        self._log_weights()
        self._l(f"\nPreparando áudio...\n", "info")
        self._set_buttons(False, False, False, False)
        self.exercise_active = True
        threading.Thread(target=self._play_thread, args=(text,), daemon=True).start()

    def _play_thread(self, text):
        try:
            audio = text_to_wav(text, self.current_freq, self.config["wpm"], self.config["farnsworth"])
            tmp = os.path.join(tempfile.gettempdir(), 'cw_ex.wav')
            save_wav(audio, tmp)
            dur = len(audio) / SAMPLE_RATE
            self.root.after(0, lambda: self._l(f"Áudio: {dur:.1f}s\n\n", "info"))
            for i in range(5, 0, -1):
                self.root.after(0, lambda n=i: self.score_var.set(str(n)))
                time.sleep(1)
            self.root.after(0, lambda: self.score_var.set("TX"))
            self.is_playing = True
            play_wav_blocking(tmp)
            self.is_playing = False
            self.root.after(0, self._after_play)
        except Exception as e:
            self.is_playing = False
            self.root.after(0, lambda: self._l(f"\nERRO: {e}\n", "err"))
            self.root.after(0, lambda: self._set_buttons(True, False, False, True))

    def _after_play(self):
        self.score_var.set("...")
        self._set_buttons(False, True, True, True)
        self.inp.focus_set()
        self._l("Transmissao finalizada. Digite e clique Corrigir.\n", "accent")

    def _on_replay(self):
        if self.is_playing or not self.current_groups:
            return
        self._set_buttons(False, False, False, False)
        text = ' '.join(self.current_groups)
        threading.Thread(target=self._replay_thread, args=(text,), daemon=True).start()

    def _replay_thread(self, text):
        try:
            audio = text_to_wav(text, self.current_freq, self.config["wpm"], self.config["farnsworth"])
            tmp = os.path.join(tempfile.gettempdir(), 'cw_ex.wav')
            save_wav(audio, tmp)
            self.is_playing = True
            self.root.after(0, lambda: self.score_var.set("TX"))
            play_wav_blocking(tmp)
            self.is_playing = False
            self.root.after(0, lambda: self.score_var.set("..."))
            self.root.after(0, lambda: self._set_buttons(False, True, True, True))
        except Exception as e:
            self.is_playing = False
            self.root.after(0, lambda: self._l(f"\nERRO: {e}\n", "err"))

    def _on_sample(self):
        if self.is_playing:
            return
        self._set_buttons(False, False, False, False)
        threading.Thread(target=self._sample_thread, daemon=True).start()

    def _sample_thread(self):
        try:
            lc = CHAR_SEQUENCE[:self.config["lesson"]]
            freq = randomize_frequency(self.config)
            self.root.after(0, self._lc)
            self.root.after(0, lambda: self._l(f"Amostra ({freq} Hz):\n\n", "accent"))
            for char in lc:
                m = MORSE_CODE.get(char, '')
                self.root.after(0, lambda c=char, mo=m: self._l(f"  {c}  ({mo})\n"))
                audio = text_to_wav(char, freq, self.config["wpm"], self.config["wpm"])
                audio.extend(generate_silence_fast(0.5))
                tmp = os.path.join(tempfile.gettempdir(), 'cw_smp.wav')
                save_wav(audio, tmp)
                self.is_playing = True
                play_wav_blocking(tmp)
                self.is_playing = False
            self.root.after(0, lambda: self._l("\nAmostra concluída.\n", "ok"))
        except Exception as e:
            self.root.after(0, lambda: self._l(f"\nERRO: {e}\n", "err"))
        finally:
            self.is_playing = False
            has_ex = bool(self.current_groups)
            self.root.after(0, lambda: self._set_buttons(True, has_ex, has_ex, True))

    def _on_check(self):
        if not self.current_groups:
            return
        received = self.inp.get("1.0", tk.END).strip().upper()
        if not received:
            messagebox.showwarning("Aviso", "Digite o que voce ouviu antes de corrigir.")
            return

        groups = self.current_groups
        lc = CHAR_SEQUENCE[:self.config["lesson"]]
        total, correct, accuracy, details, char_results = compare_groups(groups, received)

        for ch, hit in char_results:
            update_char_history(self.config, ch, hit)

        self.score_var.set(f"{accuracy:.0f}%")
        self._lc()
        self._l("RESULTADO\n\n", "bold")
        self._l("  ENVIADO | RECEBIDO\n", "accent")
        self._l("  --------  --------\n", "info")

        for sg, rg, gc in details:
            self._l(f"  {sg:6s}  |  ")
            if not rg:
                self._l("____", "err")
                self._l(f"   X   (0/{len(sg)})\n", "err")
            else:
                for ci in range(len(sg)):
                    if ci < len(rg):
                        if sg[ci] == rg[ci]:
                            self._l(rg[ci], "ok")
                        else:
                            self._l(rg[ci].lower(), "err")
                    else:
                        self._l("_", "err")
                ok = "OK" if sg == rg.upper() else "X "
                self._l(f"   {ok}  ({gc}/{len(sg)})\n", "ok" if ok == "OK" else "err")

        self._l(f"\n  Total: {total}  Corretos: {correct}  Erros: {total - correct}\n")
        bl = 30
        fl = int(bl * accuracy / 100)
        self._l(f"  [{'#' * fl + '.' * (bl - fl)}] {accuracy:.1f}%\n",
                "ok" if accuracy >= self.config["pass_threshold"] else "err")
        self._l(f"  Frequência: {self.current_freq} Hz\n\n", "info")

        # Erros
        erros = {}
        for ch, hit in char_results:
            if not hit:
                erros[ch] = erros.get(ch, 0) + 1
        if erros:
            self._l("  Erros neste exercício:\n", "err")
            for ch in sorted(erros, key=erros.get, reverse=True):
                m = MORSE_CODE.get(ch, '')
                a = get_char_accuracy(self.config, ch)
                self._l(f"    {ch} ({m}) - {erros[ch]}x  [média: {a*100:.0f}%]\n", "err")
            self._l("\n")

        # Pass/fail check
        thr = self.config["pass_threshold"] / 100.0
        win = self.config["window_size"]
        below = []
        nodata = []
        for ch in lc:
            h = self.config["char_history"].get(ch, [])
            if len(h) < win:
                nodata.append((ch, len(h)))
            elif get_char_accuracy(self.config, ch) < thr:
                below.append((ch, get_char_accuracy(self.config, ch)))

        passed = len(below) == 0 and len(nodata) == 0
        log_to_excel(self.config, lc, self.current_freq, total, correct, accuracy, passed)

        if passed:
            self._l("  >>> APROVADO! Todos os caracteres >= meta! <<<\n\n", "ok")
            if self.config["lesson"] < len(CHAR_SEQUENCE):
                self.config["lesson"] += 1
                nc = CHAR_SEQUENCE[self.config["lesson"] - 1]
                self._l(f"  Avancando para licao {self.config['lesson']}\n", "ok")
                self._l(f"  Novo caractere: {nc} ({MORSE_CODE.get(nc, '')})\n", "accent")
        else:
            self._l(f"  Repetindo licao {self.config['lesson']}.\n", "err")
            self._l(f"  Criterio: TODOS chars >= {self.config['pass_threshold']}% (janela {win})\n\n", "info")
            if below:
                below.sort(key=lambda x: x[1])
                self._l(f"  Abaixo da meta ({len(below)}):\n", "err")
                for ch, ac in below:
                    m = MORSE_CODE.get(ch, '')
                    h = self.config["char_history"].get(ch, [])
                    rc = ''.join(['o' if x else 'x' for x in h[-win:]])
                    self._l(f"    {ch} ({m:8s}) {ac*100:5.1f}%  [{rc}]\n", "err")
            if nodata:
                self._l(f"\n  Poucas amostras ({len(nodata)}):\n", "info")
                for ch, n in nodata:
                    m = MORSE_CODE.get(ch, '')
                    self._l(f"    {ch} ({m:8s}) {n}/{win}\n", "info")

        save_config(self.config)
        self.exercise_active = False
        self.current_groups = []
        self._set_buttons(True, False, False, True)
        self._update_info()
        self._refresh_stats()

    def _log_weights(self):
        lc, wts = calculate_weights(self.config)
        tw = sum(wts)
        items = sorted(zip(lc, wts), key=lambda x: x[1], reverse=True)
        self._l("\nPesos adaptativos (top 5 foco):\n", "accent")
        for ch, w in items[:5]:
            a = get_char_accuracy(self.config, ch)
            p = w / tw * 100
            h = self.config["char_history"].get(ch, [])
            astr = f"{a*100:.0f}%" if h else "--"
            tag = "err" if w > self.config["weight_ceiling"] * 0.7 else "info"
            self._l(f"  {ch}  média:{astr:>5s}  prob:{p:4.1f}%\n", tag)

    # ── Config Actions ─────────────────────────────────────

    def _save_cfg(self):
        for key, (var, typ, vmin, vmax) in self.cfg_vars.items():
            try:
                val = int(var.get()) if typ == "int" else float(var.get())
                val = max(vmin, min(val, vmax))
                self.config[key] = val
                var.set(str(val))
            except ValueError:
                pass
        self.config["farnsworth"] = min(self.config["farnsworth"], self.config["wpm"])
        self.config["weight_ceiling"] = max(self.config["weight_ceiling"], self.config["weight_floor"] + 0.1)
        save_config(self.config)
        r = self.config["weight_ceiling"] / self.config["weight_floor"]
        self.cfg_info.set(f"Salvo!  Razão ruim/bom: {r:.0f}x  |  Aprovação: TODOS chars >= {self.config['pass_threshold']}%")
        self._update_info()

    def _reset(self):
        if messagebox.askyesno("Resetar", "Resetar todas as estatísticas?"):
            self.config["char_history"] = {}
            save_config(self.config)
            self._refresh_stats()
            self._update_info()
            self.cfg_info.set("Estatísticas resetadas!")

    def _defaults(self):
        if messagebox.askyesno("Restaurar", "Restaurar configurações padrão?\n(Stats mantidas)"):
            hist = self.config.get("char_history", {})
            self.config = dict(DEFAULT_CONFIG)
            self.config["char_history"] = hist
            save_config(self.config)
            for key, (var, *_) in self.cfg_vars.items():
                var.set(str(self.config.get(key, DEFAULT_CONFIG[key])))
            self._update_info()
            self.cfg_info.set("Configuraçoes restauradas!")

    # ── Stats ──────────────────────────────────────────────

    def _refresh_stats(self):
        w = self.st_txt
        w.config(state=tk.NORMAL)
        w.delete("1.0", tk.END)
        c = self.config
        lc = CHAR_SEQUENCE[:c["lesson"]]
        win = c["window_size"]
        _, wts = calculate_weights(c)
        wm = dict(zip(lc, wts))

        data = []
        for ch in lc:
            h = c["char_history"].get(ch, [])
            n = len(h)
            a = sum(h) / n if n > 0 else -1
            data.append((ch, a, n))
        data.sort(key=lambda x: x[1])

        thr = c["pass_threshold"] / 100.0
        w.insert(tk.END, f"  Média móvel: últimas {win} aparições    |    Meta: {c['pass_threshold']}%\n\n", "header")
        w.insert(tk.END, "  CHAR  MORSE     AMOST   MÉDIA    PESO   HISTÓRICO\n", "accent")
        w.insert(tk.END, "  " + "-" * 62 + "\n", "info")

        for ch, a, n in data:
            m = MORSE_CODE.get(ch, '?')
            wt = wm.get(ch, 0.5)
            h = c["char_history"].get(ch, [])
            if n == 0:
                astr = "  --  "
                tag = "info"
            else:
                astr = f"{a*100:5.1f}%"
                tag = "ok" if a >= thr else "err"
            rc = ''.join(['o' if x else 'x' for x in h[-win:]])
            w.insert(tk.END, f"  {ch:>4s}  {m:<8s}  {n:>3d}/{win:<3d} {astr}  {wt:.2f}   {rc}\n", tag)

        cwd = [ch for ch in lc if len(c["char_history"].get(ch, [])) > 0]
        w.insert(tk.END, "\n  " + "-" * 62 + "\n", "info")
        if cwd:
            accs = [get_char_accuracy(c, ch) for ch in cwd]
            avg = sum(accs) / len(accs) * 100
            worst = min(cwd, key=lambda ch: get_char_accuracy(c, ch))
            best = max(cwd, key=lambda ch: get_char_accuracy(c, ch))
            w.insert(tk.END, f"  Média geral: {avg:.1f}%\n")
            w.insert(tk.END, f"  Melhor: {best} ({get_char_accuracy(c, best)*100:.1f}%)\n", "ok")
            w.insert(tk.END, f"  Pior:   {worst} ({get_char_accuracy(c, worst)*100:.1f}%)\n", "err")
            ready = sum(1 for ch in lc
                        if len(c["char_history"].get(ch, [])) >= win
                        and get_char_accuracy(c, ch) >= thr)
            w.insert(tk.END, f"\n  Progresso: {ready}/{len(lc)} chars aprovados\n", "accent")
        else:
            w.insert(tk.END, "  Sem dados ainda.\n", "info")
        w.config(state=tk.DISABLED)


# ===========================================================
#  MAIN
# ===========================================================

def main():
    root = tk.Tk()
    try:
        import ctypes
        root.update()
        hwnd = ctypes.windll.user32.GetParent(root.winfo_id())
        ctypes.windll.dwmapi.DwmSetWindowAttribute(
            hwnd, 20, ctypes.byref(ctypes.c_int(1)), ctypes.sizeof(ctypes.c_int))
    except Exception:
        pass

    app = CWTrainerApp(root)

    def on_close():
        if app.is_playing:
            stop_audio()
        save_config(app.config)
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()


if __name__ == '__main__':
    main()
