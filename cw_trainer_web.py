#!/usr/bin/env python3
"""
+===========================================================+
|     CW ADAPTIVE TRAINER  v7  (Streamlit)                   |
|     Treino de CW por PY2TAE (Lucas)                        |
|                                                            |
|  Roda no navegador: celular, tablet, desktop, qualquer OS  |
|  Audio via browser = funciona em tudo                      |
|                                                            |
|  Uso: streamlit run cw_trainer_web.py                      |
|  Deps: pip install streamlit openpyxl                      |
+===========================================================+
"""

import os, sys, json, math, wave, array, random, io, datetime
import streamlit as st

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    st.error("openpyxl necessario: pip install openpyxl")
    st.stop()


# ===========================================================
#  CONSTANTES
# ===========================================================

CHAR_SEQUENCE = list("ETIANMSURWDKGOHVFLPJBXCYZQ0123456789.,?/=")

MORSE_CODE = {
    'A': '.-',    'B': '-...',  'C': '-.-.',  'D': '-..',   'E': '.',
    'F': '..-.',  'G': '--.',   'H': '....',  'I': '..',    'J': '.---',
    'K': '-.-',   'L': '.-..',  'M': '--',    'N': '-.',    'O': '---',
    'P': '.--.',  'Q': '--.-',  'R': '.-.',   'S': '...',   'T': '-',
    'U': '..-',   'V': '...-',  'W': '.--',   'X': '-..-',  'Y': '-.--',
    'Z': '--..',
    '0': '-----', '1': '.----', '2': '..---', '3': '...--', '4': '....-',
    '5': '.....', '6': '-....', '7': '--...', '8': '---..', '9': '----.',
    '.': '.-.-.-', ',': '--..--', '?': '..--..', '/': '-..-.', '=': '-...-',
}

SAMPLE_RATE = 22050
RAMP_MS = 3
TWO_PI = 2.0 * math.pi

DATA_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(DATA_DIR, "cw_config.json")
EXCEL_FILE = os.path.join(DATA_DIR, "cw_telemetria.xlsx")


# ===========================================================
#  AUDIO ENGINE (WAV em memoria para o browser)
# ===========================================================

def _tone(duration_s, frequency):
    n = int(SAMPLE_RATE * duration_s)
    if n == 0:
        return array.array('h')
    ramp = int(SAMPLE_RATE * RAMP_MS / 1000)
    buf = array.array('h', [0] * n)
    for i in range(n):
        val = math.sin(TWO_PI * frequency * i / SAMPLE_RATE)
        if i < ramp:
            val *= i / ramp
        elif i >= n - ramp:
            val *= (n - 1 - i) / ramp
        buf[i] = int(val * 24000)
    return buf


def _silence(duration_s):
    return array.array('h', [0] * int(SAMPLE_RATE * duration_s))


def _timing(wpm, farn):
    dit = 1.2 / wpm
    dah = 3 * dit
    intra = dit
    if farn < wpm:
        delta = (60.0 / farn - 50 * dit) / 19.0
        inter_c = 3 * delta
        inter_w = 7 * delta
    else:
        inter_c = 3 * dit
        inter_w = 7 * dit
    return dit, dah, intra, inter_c, inter_w


def _array_to_wav_bytes(audio_arr):
    buf = io.BytesIO()
    with wave.open(buf, 'w') as wf:
        wf.setnchannels(1)
        wf.setsampwidth(2)
        wf.setframerate(SAMPLE_RATE)
        wf.writeframes(audio_arr.tobytes())
    buf.seek(0)
    return buf


def text_to_audio(text, frequency, wpm, farn):
    """Retorna (BytesIO wav, duracao_s)."""
    dit_t, dah_t, intra_t, ic_t, iw_t = _timing(wpm, farn)
    td = _tone(dit_t, frequency)
    ta = _tone(dah_t, frequency)
    si = _silence(intra_t)
    sc = _silence(ic_t)
    sw = _silence(iw_t)

    audio = array.array('h')
    for i, ch in enumerate(text):
        if ch == ' ':
            audio.extend(sw)
            continue
        morse = MORSE_CODE.get(ch.upper(), '')
        if not morse:
            continue
        for j, sym in enumerate(morse):
            audio.extend(td if sym == '.' else ta)
            if j < len(morse) - 1:
                audio.extend(si)
        if i < len(text) - 1 and text[i + 1] != ' ':
            audio.extend(sc)

    return _array_to_wav_bytes(audio), len(audio) / SAMPLE_RATE


def sample_audio(chars, frequency, wpm):
    """Gera WAV com cada caractere + pausa."""
    dit_t, dah_t, intra_t, _, _ = _timing(wpm, wpm)
    td = _tone(dit_t, frequency)
    ta = _tone(dah_t, frequency)
    si = _silence(intra_t)
    pause = _silence(0.6)

    audio = array.array('h')
    for ch in chars:
        morse = MORSE_CODE.get(ch.upper(), '')
        if not morse:
            continue
        for j, sym in enumerate(morse):
            audio.extend(td if sym == '.' else ta)
            if j < len(morse) - 1:
                audio.extend(si)
        audio.extend(pause)

    return _array_to_wav_bytes(audio)


# ===========================================================
#  CONFIG & PERSISTENCE
# ===========================================================

DEFAULT_CONFIG = {
    "frequency": 700, "freq_variation": 0, "wpm": 10, "farnsworth": 6,
    "lesson": 2, "num_groups": 6, "group_size": 4, "pass_threshold": 90,
    "window_size": 10, "weight_floor": 0.1, "weight_ceiling": 1.0,
    "char_history": {},
}

def load_config():
    cfg = dict(DEFAULT_CONFIG)
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                cfg.update(json.load(f))
        except (json.JSONDecodeError, IOError):
            pass
    if not isinstance(cfg.get("char_history"), dict):
        cfg["char_history"] = {}
    return cfg

def save_config(cfg):
    with open(CONFIG_FILE, 'w') as f:
        json.dump(cfg, f, indent=2)

def char_acc(cfg, ch):
    h = cfg["char_history"].get(ch, [])
    return sum(h) / len(h) if h else 0.5

def update_hist(cfg, ch, hit):
    win = cfg.get("window_size", 10)
    if ch not in cfg["char_history"]:
        cfg["char_history"][ch] = []
    cfg["char_history"][ch].append(1 if hit else 0)
    if len(cfg["char_history"][ch]) > win:
        cfg["char_history"][ch] = cfg["char_history"][ch][-win:]


# ===========================================================
#  WEIGHTED GROUPS
# ===========================================================

def calc_weights(cfg):
    fl = cfg.get("weight_floor", 0.1)
    ce = cfg.get("weight_ceiling", 1.0)
    lc = CHAR_SEQUENCE[:cfg["lesson"]]
    return lc, [fl + (ce - fl) * (1.0 - char_acc(cfg, c)) for c in lc]

def gen_groups(cfg):
    lc, wts = calc_weights(cfg)
    return [''.join(random.choices(lc, weights=wts, k=cfg["group_size"]))
            for _ in range(cfg["num_groups"])]

def rand_freq(cfg):
    v = cfg.get("freq_variation", 0)
    return cfg["frequency"] + (random.randint(-v, v) if v else 0)


# ===========================================================
#  COMPARISON
# ===========================================================

def compare(sent, recv_text):
    rg = recv_text.upper().split()
    total = correct = 0
    details, cres = [], []
    for i, sg in enumerate(sent):
        r = rg[i] if i < len(rg) else ''
        gc = 0
        for ci in range(len(sg)):
            total += 1
            hit = ci < len(r) and sg[ci] == r[ci]
            if hit: gc += 1
            cres.append((sg[ci], hit))
        correct += gc
        details.append((sg, r, gc))
    return total, correct, (correct/total*100 if total else 0), details, cres


# ===========================================================
#  EXCEL
# ===========================================================

def init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetria CW"
    headers = ["Data","Hora","Licao","Caracteres","WPM","Farnsworth",
               "Freq (Hz)","Grupos","Chars/Grp","Total","Corretos","Acerto (%)","Resultado"]
    hf = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hfn = Font(bold=True, color="FFFFFF", size=11)
    tb = Border(left=Side(style='thin'),right=Side(style='thin'),
                top=Side(style='thin'),bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hf, hfn, Alignment(horizontal='center'), tb
    wb.save(EXCEL_FILE)

def log_excel(cfg, lc, freq, total, correct, acc, passed):
    """Registra tentativa no disco (local) e no session_state (cloud)."""
    now = datetime.datetime.now()
    res = "APROVADO" if passed else "REPETIR"
    row = [now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), cfg["lesson"],
           ' '.join(lc), cfg["wpm"], cfg["farnsworth"], freq,
           cfg["num_groups"], cfg["group_size"], total, correct, round(acc,1), res]

    # Salvar em session_state (sempre funciona, cloud ou local)
    if "telemetry" not in st.session_state:
        st.session_state.telemetry = []
    st.session_state.telemetry.append(row)

    # Tentar salvar no disco também (funciona local, pode falhar no cloud)
    try:
        init_excel()
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        tb = Border(left=Side(style='thin'),right=Side(style='thin'),
                    top=Side(style='thin'),bottom=Side(style='thin'))
        ws.append(row)
        rn = ws.max_row
        for col in range(1, len(row)+1):
            c = ws.cell(row=rn, column=col)
            c.border, c.alignment = tb, Alignment(horizontal='center')
        rc = ws.cell(row=rn, column=len(row))
        if res == "APROVADO":
            rc.fill = PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
            rc.font = Font(color="006100", bold=True)
        else:
            rc.fill = PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid")
            rc.font = Font(color="9C0006", bold=True)
        wb.save(EXCEL_FILE)
    except Exception:
        pass  # No cloud, pode falhar — telemetria fica no session_state


def _build_telemetry_excel():
    """Gera Excel da telemetria em memória para download."""
    rows = st.session_state.get("telemetry", [])
    if not rows:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Telemetria CW"
    headers = ["Data","Hora","Licao","Caracteres","WPM","Farnsworth",
               "Freq (Hz)","Grupos","Chars/Grp","Total","Corretos","Acerto (%)","Resultado"]
    hf = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    hfn = Font(bold=True, color="FFFFFF", size=11)
    tb = Border(left=Side(style='thin'),right=Side(style='thin'),
                top=Side(style='thin'),bottom=Side(style='thin'))
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = hf, hfn, Alignment(horizontal='center'), tb

    for row in rows:
        ws.append(row)
        rn = ws.max_row
        for col in range(1, len(row)+1):
            c = ws.cell(row=rn, column=col)
            c.border, c.alignment = tb, Alignment(horizontal='center')
        rc = ws.cell(row=rn, column=len(row))
        res = row[-1]
        if res == "APROVADO":
            rc.fill = PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
            rc.font = Font(color="006100", bold=True)
        else:
            rc.fill = PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type="solid")
            rc.font = Font(color="9C0006", bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ===========================================================
#  STREAMLIT APP
# ===========================================================

def init_state():
    for k, v in {"cfg": None, "groups": [], "freq": 0, "wav": None,
                 "dur": 0, "checked": False, "result": None,
                 "telemetry": []}.items():
        if k not in st.session_state:
            st.session_state[k] = v
    if st.session_state.cfg is None:
        # Tenta carregar do disco (uso local), senao usa defaults
        st.session_state.cfg = load_config()
        init_excel()


def sidebar():
    cfg = st.session_state.cfg
    lc = CHAR_SEQUENCE[:cfg["lesson"]]

    st.sidebar.markdown("## ⚡ CW Trainer")
    st.sidebar.caption("por PY2TAE (Lucas)")
    st.sidebar.markdown(f"**Licao {cfg['lesson']}/{len(CHAR_SEQUENCE)}**")
    st.sidebar.code(' '.join(lc), language=None)
    if cfg["lesson"] < len(CHAR_SEQUENCE):
        nx = CHAR_SEQUENCE[cfg["lesson"]]
        st.sidebar.markdown(f"Proximo: **{nx}** (`{MORSE_CODE.get(nx,'')}`)")

    st.sidebar.divider()

    with st.sidebar.expander("⚙️ Configuracoes"):
        cfg["lesson"] = st.number_input("Licao", 2, len(CHAR_SEQUENCE), cfg["lesson"], key="sl")
        cfg["frequency"] = st.number_input("Freq base (Hz)", 200, 1500, cfg["frequency"], key="sf")
        cfg["freq_variation"] = st.number_input("Variacao +/- Hz", 0, 200, cfg["freq_variation"], key="sv")
        cfg["wpm"] = st.number_input("WPM", 5, 60, cfg["wpm"], key="sw")
        cfg["farnsworth"] = st.number_input("Farnsworth", 2, cfg["wpm"], min(cfg["farnsworth"], cfg["wpm"]), key="sn")
        cfg["num_groups"] = st.number_input("Grupos", 1, 50, cfg["num_groups"], key="sg")
        cfg["group_size"] = st.number_input("Chars/grupo", 2, 10, cfg["group_size"], key="sc")
        cfg["pass_threshold"] = st.number_input("Meta (%)", 50, 100, cfg["pass_threshold"], key="st_")
        cfg["window_size"] = st.number_input("Janela", 3, 100, cfg["window_size"], key="sj")
        cfg["weight_floor"] = st.number_input("Peso min", 0.01, 0.5, cfg["weight_floor"], format="%.2f", key="spf")
        cfg["weight_ceiling"] = st.number_input("Peso max", 0.1, 10.0, cfg["weight_ceiling"], format="%.1f", key="spc")

        if st.button("💾 Salvar"):
            cfg["weight_ceiling"] = max(cfg["weight_ceiling"], cfg["weight_floor"]+0.1)
            save_config(cfg)
            st.success("Salvo!")

    with st.sidebar.expander("🔧 Acoes"):
        if st.button("Resetar estatisticas"):
            cfg["char_history"] = {}
            save_config(cfg)
            st.rerun()
        if st.button("Restaurar padroes"):
            h = cfg.get("char_history", {})
            st.session_state.cfg = dict(DEFAULT_CONFIG)
            st.session_state.cfg["char_history"] = h
            save_config(st.session_state.cfg)
            st.rerun()

    with st.sidebar.expander("💾 Progresso (salvar/carregar)"):
        st.caption("Seu progresso fica na sessao do navegador. "
                   "Baixe o arquivo para nao perder!")

        # ── Download do progresso ──
        progress_data = json.dumps(cfg, indent=2, ensure_ascii=False)
        st.download_button(
            label="⬇️ Baixar meu progresso",
            data=progress_data,
            file_name="cw_progresso.json",
            mime="application/json",
            use_container_width=True,
        )

        # ── Upload do progresso ──
        uploaded = st.file_uploader("⬆️ Carregar progresso", type=["json"],
                                     label_visibility="collapsed")
        if uploaded is not None:
            try:
                loaded = json.loads(uploaded.getvalue().decode("utf-8"))
                # Validar que tem os campos essenciais
                if "lesson" in loaded and "char_history" in loaded:
                    # Mesclar com defaults para garantir campos novos
                    merged = dict(DEFAULT_CONFIG)
                    merged.update(loaded)
                    st.session_state.cfg = merged
                    save_config(merged)
                    st.success(f"Progresso carregado! Licao {merged['lesson']}, "
                              f"{len(merged.get('char_history', {}))} chars com dados.")
                    st.rerun()
                else:
                    st.error("Arquivo invalido. Use o JSON baixado pelo botao acima.")
            except (json.JSONDecodeError, UnicodeDecodeError):
                st.error("Arquivo corrompido ou formato invalido.")

        # ── Download da telemetria Excel ──
        if st.session_state.telemetry:
            tel_wb = _build_telemetry_excel()
            if tel_wb:
                st.download_button(
                    label="📊 Baixar telemetria (.xlsx)",
                    data=tel_wb,
                    file_name="cw_telemetria.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    st.sidebar.divider()
    r = cfg["weight_ceiling"] / cfg["weight_floor"] if cfg["weight_floor"] > 0 else 999
    st.sidebar.caption(
        f"{cfg['wpm']} WPM | Farns: {cfg['farnsworth']} | "
        f"{cfg['frequency']}Hz±{cfg['freq_variation']}\n\n"
        f"{cfg['num_groups']}x{cfg['group_size']} | Meta: {cfg['pass_threshold']}% | Razao: {r:.0f}x"
    )


def tab_exercise():
    cfg = st.session_state.cfg
    c1, c2, c3 = st.columns([1, 1, 1])

    with c1:
        play = st.button("▶  GERAR E TOCAR", type="primary", use_container_width=True)
    with c2:
        samp = st.button("♪  Amostra", use_container_width=True)
    with c3:
        chk = st.button("✓  Corrigir", type="primary", use_container_width=True)

    # PLAY
    if play:
        st.session_state.groups = gen_groups(cfg)
        st.session_state.freq = rand_freq(cfg)
        st.session_state.checked = False
        st.session_state.result = None
        text = ' '.join(st.session_state.groups)
        with st.spinner("Gerando audio..."):
            buf, dur = text_to_audio(text, st.session_state.freq, cfg["wpm"], cfg["farnsworth"])
            st.session_state.wav = buf.getvalue()
            st.session_state.dur = dur

    # SAMPLE
    if samp:
        lc = CHAR_SEQUENCE[:cfg["lesson"]]
        freq = rand_freq(cfg)
        with st.spinner("Gerando amostra..."):
            buf = sample_audio(lc, freq, cfg["wpm"])
        st.audio(buf, format="audio/wav", autoplay=True)
        cols = st.columns(min(len(lc), 8))
        for i, ch in enumerate(lc):
            cols[i % len(cols)].code(f"{ch} {MORSE_CODE.get(ch,'')}")
        return

    # PLAYER
    if st.session_state.wav and not st.session_state.checked:
        st.divider()
        groups = st.session_state.groups
        tc = sum(len(g) for g in groups)

        with st.expander(f"ℹ️ {len(groups)} grupos x {cfg['group_size']} = {tc} chars | {st.session_state.freq} Hz"):
            lc, wts = calc_weights(cfg)
            tw = sum(wts)
            items = sorted(zip(lc, wts), key=lambda x: x[1], reverse=True)[:5]
            for ch, w in items:
                a = char_acc(cfg, ch)
                h = cfg["char_history"].get(ch, [])
                astr = f"{a*100:.0f}%" if h else "--"
                foco = " 🎯" if w > cfg["weight_ceiling"] * 0.7 else ""
                st.text(f"  {ch}  media: {astr:>5s}  prob: {w/tw*100:4.1f}%{foco}")

        st.audio(st.session_state.wav, format="audio/wav", autoplay=True)
        st.caption(f"⏱️ {st.session_state.dur:.1f}s — Aperte play no player acima e digite abaixo")

    # INPUT
    if st.session_state.groups and not st.session_state.checked:
        st.markdown("### ✏️ Digite o que ouviu")
        st.caption("Separe os grupos com espaco")
        recv = st.text_input("Resposta:", key="ans", placeholder="ETIA MSNR ...",
                             label_visibility="collapsed")

        if chk:
            if recv and recv.strip():
                do_check(recv.strip().upper())
            else:
                st.warning("Digite algo antes de corrigir.")

    # RESULT
    if st.session_state.result:
        show_result()


def do_check(recv):
    cfg = st.session_state.cfg
    groups = st.session_state.groups
    lc = CHAR_SEQUENCE[:cfg["lesson"]]

    total, correct, acc, details, cres = compare(groups, recv)

    for ch, hit in cres:
        update_hist(cfg, ch, hit)

    thr = cfg["pass_threshold"] / 100.0
    win = cfg["window_size"]
    below, nodata = [], []
    for ch in lc:
        h = cfg["char_history"].get(ch, [])
        if len(h) < win:
            nodata.append((ch, len(h)))
        elif char_acc(cfg, ch) < thr:
            below.append((ch, char_acc(cfg, ch)))

    passed = not below and not nodata

    log_excel(cfg, lc, st.session_state.freq, total, correct, acc, passed)

    if passed and cfg["lesson"] < len(CHAR_SEQUENCE):
        cfg["lesson"] += 1

    save_config(cfg)
    st.session_state.result = {
        "total": total, "correct": correct, "acc": acc,
        "details": details, "cres": cres,
        "passed": passed, "below": below, "nodata": nodata,
        "freq": st.session_state.freq, "recv": recv,
    }
    st.session_state.checked = True


def show_result():
    cfg = st.session_state.cfg
    r = st.session_state.result

    st.divider()
    c1, c2, c3 = st.columns(3)
    c1.metric("Acerto", f"{r['acc']:.1f}%")
    c2.metric("Corretos", f"{r['correct']}/{r['total']}")
    c3.metric("Freq", f"{r['freq']} Hz")
    st.progress(min(r['acc'] / 100, 1.0))

    # Comparacao
    st.markdown("#### Comparacao grupo a grupo")
    for sg, rg, gc in r["details"]:
        if not rg:
            marked = " ".join(["🔴"] * len(sg))
        else:
            parts = []
            for ci in range(len(sg)):
                if ci < len(rg):
                    if sg[ci] == rg[ci]:
                        parts.append(f"**{rg[ci]}**")
                    else:
                        parts.append(f"~~{rg[ci].lower()}~~")
                else:
                    parts.append("⬜")
            marked = " ".join(parts)
        icon = "✅" if sg == (rg.upper() if rg else '') else "❌"
        st.markdown(f"`{sg}` → {marked}  {icon} ({gc}/{len(sg)})")

    # Erros
    erros = {}
    for ch, hit in r["cres"]:
        if not hit:
            erros[ch] = erros.get(ch, 0) + 1
    if erros:
        st.markdown("#### Erros neste exercicio")
        for ch in sorted(erros, key=erros.get, reverse=True):
            m = MORSE_CODE.get(ch, '')
            a = char_acc(cfg, ch)
            st.markdown(f"- **{ch}** (`{m}`) — {erros[ch]}x erro — media: {a*100:.0f}%")

    # Pass/fail
    st.divider()
    if r["passed"]:
        st.success(f"✅ APROVADO! Todos chars >= {cfg['pass_threshold']}%")
        if cfg["lesson"] <= len(CHAR_SEQUENCE):
            nc = CHAR_SEQUENCE[cfg["lesson"] - 1]
            st.info(f"Avancando para licao {cfg['lesson']} — **{nc}** (`{MORSE_CODE.get(nc,'')}`)")
    else:
        st.error(f"Repetindo licao {cfg['lesson']}. TODOS chars >= {cfg['pass_threshold']}% (janela {cfg['window_size']})")
        if r["below"]:
            st.markdown("**Abaixo da meta:**")
            for ch, ac in sorted(r["below"], key=lambda x: x[1]):
                m = MORSE_CODE.get(ch, '')
                h = cfg["char_history"].get(ch, [])
                vis = ''.join(['🟢' if x else '🔴' for x in h[-cfg["window_size"]:]])
                st.markdown(f"- **{ch}** (`{m}`) — {ac*100:.1f}% — {vis}")
        if r["nodata"]:
            st.markdown("**Poucas amostras:**")
            for ch, n in r["nodata"]:
                st.markdown(f"- **{ch}** (`{MORSE_CODE.get(ch,'')}`) — {n}/{cfg['window_size']}")

    if st.button("🔄 Novo exercicio", type="primary"):
        st.session_state.groups = []
        st.session_state.wav = None
        st.session_state.checked = False
        st.session_state.result = None
        st.rerun()


def tab_stats():
    cfg = st.session_state.cfg
    lc = CHAR_SEQUENCE[:cfg["lesson"]]
    win = cfg["window_size"]
    thr = cfg["pass_threshold"] / 100.0
    _, wts = calc_weights(cfg)
    wm = dict(zip(lc, wts))

    ready = sum(1 for ch in lc
                if len(cfg["char_history"].get(ch, [])) >= win and char_acc(cfg, ch) >= thr)
    st.progress(ready / len(lc) if lc else 0)
    st.caption(f"Progresso: {ready}/{len(lc)} chars aprovados (>= {cfg['pass_threshold']}%)")

    data = []
    for ch in lc:
        h = cfg["char_history"].get(ch, [])
        n = len(h)
        acc = sum(h)/n if n > 0 else -1
        vis = ''.join(['🟢' if x else '🔴' for x in h[-win:]])
        ok = "✅" if n >= win and acc >= thr else ("⏳" if n < win else "❌")
        data.append({"": ok, "Char": ch, "Morse": MORSE_CODE.get(ch,'?'),
                     "N": f"{n}/{win}", "Media": f"{acc*100:.1f}%" if n else "--",
                     "Peso": f"{wm.get(ch,0.5):.2f}", "Hist": vis or "—"})
    data.sort(key=lambda x: float(x["Media"].replace('%','').replace('--','-1')))
    st.dataframe(data, use_container_width=True, hide_index=True,
                 height=min(35*len(data)+38, 600))

    cwd = [ch for ch in lc if cfg["char_history"].get(ch)]
    if cwd:
        accs = [char_acc(cfg, ch) for ch in cwd]
        avg = sum(accs)/len(accs)*100
        worst = min(cwd, key=lambda c: char_acc(cfg, c))
        best = max(cwd, key=lambda c: char_acc(cfg, c))
        c1, c2, c3 = st.columns(3)
        c1.metric("Media", f"{avg:.1f}%")
        c2.metric(f"Melhor: {best}", f"{char_acc(cfg,best)*100:.1f}%")
        c3.metric(f"Pior: {worst}", f"{char_acc(cfg,worst)*100:.1f}%")


def tab_ref():
    cfg = st.session_state.cfg
    lc = CHAR_SEQUENCE[:cfg["lesson"]]
    st.markdown("### Tabela Morse")
    cols = st.columns(min(len(lc), 6))
    for i, ch in enumerate(lc):
        m = MORSE_CODE.get(ch, '')
        a = char_acc(cfg, ch)
        h = cfg["char_history"].get(ch, [])
        astr = f"{a*100:.0f}%" if h else "--"
        cols[i % len(cols)].code(f"{ch}  {m:<8s} {astr}")


# ===========================================================
#  MAIN
# ===========================================================

def main():
    st.set_page_config(page_title="CW Trainer - PY2TAE", page_icon="⚡",
                       layout="wide", initial_sidebar_state="expanded")
    st.markdown("""<style>
    .stApp { background-color: #1a1a2e; }
    div[data-testid="stMetric"] { background: #16213e; padding: 10px; border-radius: 8px; }
    .stTextInput input { font-family: 'Consolas','Courier New',monospace; font-size: 1.2rem; }
    </style>""", unsafe_allow_html=True)

    init_state()
    sidebar()

    t1, t2, t3 = st.tabs(["⚡ Exercicio", "📊 Estatisticas", "📖 Morse"])
    with t1: tab_exercise()
    with t2: tab_stats()
    with t3: tab_ref()


if __name__ == '__main__':
    main()
